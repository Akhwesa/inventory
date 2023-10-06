from .models import Desktop1, Laptop1
from django.http import HttpResponse
from openpyxl import Workbook
from django.contrib.auth.models import User

def desktop_excel(request):
    items = Desktop1.objects.all()    
    workstation = []
    for item in items:
        wk_name =  item.Computer_Name
        wk_make = item.Desktop_Make
        wk_SN = item.Serial_Number
        processor =  item.Processor
        os =  item.Operating_System
        office =  item.Office_Suite
        AV = item.anti_virus
        Region = item.region
        dpt = item.department
        st =  item.station
        location = item.Location
        cd = item.Condition
        assigned =  item.assign
        time =  item.last_updated
        
        hard_disks = item.hard_disks.all()
        total_disk_size = sum([hard_disks.Hard_disk_Size for hard_disks in hard_disks])

        rams = item.rams.all()
        total_ram_size = sum([rams.RAM_Size for rams in rams])
        monitor_sn = None
        if hasattr(item, 'monitor'):  # Check if Monitor object exists
            monitor_sn = item.monitor.monitor_SN
            monitor_make = item.monitor.monitor_make

        keyboard_sn = None
        if hasattr(item, 'keyboard'):  # Check if keyboard object exists
            keyboard_sn = item.keyboard.keyboard_SN

        mouse_sn = None
        if hasattr(item, 'mouse'):  # Check if mouse object exists
            mouse_sn = item.mouse.mouse_SN
        
        workstation.append({
            'hard_disk':total_disk_size,
            'ram': total_ram_size,
            'name': wk_name,
            'make': wk_make,
            'sn': wk_SN,
            'processor':processor,
            'os':os,
            'office': office,
            'av': AV,
            'region':Region,
            'department':dpt,
            'station': st,
            'location':location,
            'condition': cd,
            'user': assigned,
            "time":time,
            'monitor_sn': monitor_sn,
            'keyboard_sn': keyboard_sn, 
            'mouse_sn': mouse_sn,
            'monitor_make':monitor_make,  
            'pk':item.pk      
            
        })

    # Create a new workbook
    workbook = Workbook()

    # Get the active worksheet
    worksheet = workbook.active

    # Define the column headers
    worksheet['A1'] = 'MAKE/MODEL'
    worksheet['B1'] = 'COMPUTER NAME'
    worksheet['C1'] = 'CPU SERIAL NUMBER'
    worksheet['D1'] = 'MONITOR MAKE'
    worksheet['E1'] = 'MONITOR SERIAL NUMBER'
    worksheet['F1'] = 'RAM'
    worksheet['G1'] = 'HARDISK'
    worksheet['H1'] = 'PROCESSOR'
    worksheet['I1'] = 'OPERATING SYSTEM'
    worksheet['J1'] = 'OFFICE SUITE'
    worksheet['K1'] = 'ANTIVIRUS'
    worksheet['L1'] = 'EXACT LOCATION'
    worksheet['M1'] = 'AUTHORIZED USER'
    worksheet['N1'] = 'DEPARTMENT'
    worksheet['O1'] = 'Condition'

    # Fill in the rows with data
    for i, item in enumerate(workstation):
        assigned_user = item['user'].username if item['user'] else None
        worksheet.cell(row=i+2, column=1, value=item['make'])
        worksheet.cell(row=i+2, column=2, value=item['name'])
        worksheet.cell(row=i+2, column=3, value=item['sn'])
        worksheet.cell(row=i+2, column=4, value=item['monitor_make'])
        worksheet.cell(row=i+2, column=5, value=item['monitor_sn'])
        worksheet.cell(row=i+2, column=6, value=item['ram'])
        worksheet.cell(row=i+2, column=7, value=item['hard_disk'])
        worksheet.cell(row=i+2, column=8, value=item['processor'])
        worksheet.cell(row=i+2, column=9, value=item['os'])
        worksheet.cell(row=i+2, column=10, value=item['office'])
        worksheet.cell(row=i+2, column=11, value=item['av'])
        worksheet.cell(row=i+2, column=12, value=item['location'])
        worksheet.cell(row=i+2, column=13, value=assigned_user)
        worksheet.cell(row=i+2, column=14, value=item['department'])
        worksheet.cell(row=i+2, column=15, value=item['condition'])


    # Create a response object with the workbook as the content
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="Workstations.xlsx"'
    workbook.save(response)

    return response

def laptop_excel(request):

    items = Laptop1.objects.all()
    lappy = []
    for item in items:
        wk_name =  item.Computer_Name
        wk_make = item.Make
        wk_SN = item.Serial_Number
        processor =  item.Processor
        os =  item.Operating_System
        office =  item.Office_Suite
        AV = item.anti_virus
        Region = item.region
        dpt = item.department
        st =  item.station
        location = item.Location
        cd = item.Condition
        assigned =  item.assign
        time =  item.last_updated
        
        lt_hdds = item.lt_hdds.all()
        total_disk_size = sum([lt_hdds.LT_Hard_disk_Size for lt_hdds in lt_hdds])

        lt_rams = item.lt_rams.all()
        total_ram_size = sum([lt_rams.LT_RAM_Size for lt_rams in lt_rams])
        
        lappy.append({
            'hard_disk':total_disk_size,
            'ram': total_ram_size,
            'name': wk_name,
            'make': wk_make,
            'sn': wk_SN,
            'processor':processor,
            'os':os,
            'office': office,
            'av': AV,
            'region':Region,
            'department':dpt,
            'station': st,
            'location':location,
            'condition': cd,
            'user': assigned,
            'time':time,  
            'pk': item.pk  
            
        })    
    # Create a new workbook
    workbook = Workbook()

    # Get the active worksheet
    worksheet = workbook.active

    # Define the column headers
    worksheet['A1'] = 'MAKE/MODEL'
    worksheet['B1'] = 'COMPUTER NAME'
    worksheet['C1'] = 'CPU SERIAL NUMBER'
    worksheet['D1'] = 'RAM'
    worksheet['E1'] = 'HARDISK'
    worksheet['F1'] = 'PROCESSOR'
    worksheet['G1'] = 'OPERATING SYSTEM'
    worksheet['H1'] = 'OFFICE SUITE'
    worksheet['I1'] = 'ANTIVIRUS'
    worksheet['J1'] = 'EXACT LOCATION'
    worksheet['K1'] = 'AUTHORIZED USER'
    worksheet['L1'] = 'DEPARTMENT'
    worksheet['M1'] = 'Condition'

    # Fill in the rows with data
    for i, item in enumerate(lappy):
        assigned_user = item['user'].username if item['user'] else None
        worksheet.cell(row=i+2, column=1, value=item['make'])
        worksheet.cell(row=i+2, column=2, value=item['name'])
        worksheet.cell(row=i+2, column=3, value=item['sn'])
        worksheet.cell(row=i+2, column=4, value=item['ram'])
        worksheet.cell(row=i+2, column=5, value=item['hard_disk'])
        worksheet.cell(row=i+2, column=6, value=item['processor'])
        worksheet.cell(row=i+2, column=7, value=item['os'])
        worksheet.cell(row=i+2, column=8, value=item['office'])
        worksheet.cell(row=i+2, column=9, value=item['av'])
        worksheet.cell(row=i+2, column=10, value=item['location'])
        worksheet.cell(row=i+2, column=11, value=assigned_user)
        worksheet.cell(row=i+2, column=12, value=item['department'])
        worksheet.cell(row=i+2, column=13, value=item['condition'])


    # Create a response object with the workbook as the content
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="laptops.xlsx"'
    workbook.save(response)    

    return response